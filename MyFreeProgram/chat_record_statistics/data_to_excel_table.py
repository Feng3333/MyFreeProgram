#
# Description: Converting chat_record_info data to excel table
# Create: 2025-01-08
#
import os.path

import pandas
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def init_chat_record_table():
    chat_record_table = dict()
    chat_record_table['工号'] = list()
    chat_record_table['姓名'] = list()
    chat_record_table['发言次数'] = list()
    return chat_record_table


def write_chat_record_data_to_table(chat_record_info: dict, chat_record_table: dict):
    for key, value in chat_record_info.items():
        chat_record_table['工号'].append(key)
        person_info_list = value[0]
        chat_record_table['姓名'].append(person_info_list['name'])
        chat_record_table['发言次数'].append(person_info_list['chat_frequency'])
    return


def is_chinese(word):
    pattern = re.compile(r'[^\u4e00-\u9fa5]')
    if pattern.search(word):
        return False
    return True


# 调整excel表格中的行距
def adaptive_excel_col_width(file_path):
    wb = load_workbook(file_path)
    # 遍历excel表格中的每个sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        each_col_width = []
        # 遍历sheet中的所有列
        for i in range(1, ws.max_column + 1):
            # 找到每行的最大宽度
            max_width = 0.0
            for j in range(1, ws.max_row + 1):
                sheet_value = ws.cell(row=j, column=i).value
                sheet_value_list = [ch for ch in str(sheet_value)]
                cur_width = 0
                for ch in sheet_value_list:
                    if is_chinese(ch):
                        cur_width += 2.2
                    else:
                        cur_width += 1.1
                max_width = max(max_width, cur_width)
            each_col_width.append(max_width)
        for i in range(1, ws.max_column + 1):
            k = get_column_letter(i)
            ws.column_dimensions[k].width = min(each_col_width[i - 1], 20) + 2
    wb.save(file_path)
    wb.close()
    return


def chat_record_data_to_excel_process(chat_record_data, out_path):
    # 1. init chat_record_table
    chat_record_table = init_chat_record_table()

    # 2. generate chat_record_table from chat_record_info which is processed
    write_chat_record_data_to_table(chat_record_data, chat_record_table)

    # 3. generate origin excel from chat_record_table
    data_frame = pandas.DataFrame(chat_record_table)
    output_excel_file = os.path.join(out_path, 'Chat_History.xlsx')
    data_frame.to_excel(output_excel_file, index=False)

    # 4. 调整excel表格中的行距
    adaptive_excel_col_width(output_excel_file)
    return
